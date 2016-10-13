#! /usr/bin/env python

import sys
import re
import collections
import numpy as np
from openpyxl.reader.excel import Workbook
from openpyxl.style import Border,Style,Color, Fill
from openpyxl.cell import Cell
from datetime import datetime, timedelta
from attask_libs import * 
from attask_api import StreamClient, ObjCode, AtTaskObject

sys.path.insert(0,'/home/analytics/analytics_sandbox/python_libs');
from common_libs import *

# Logging
import log_libs as log
LOG = log.init_logging()

def project_fields():

	project=[]
	project.append('actualCompletionDate')
	project.append('actualDurationMinutes')
	project.append('actualLaborCost')
	project.append('actualStartDate')
	project.append('actualWorkRequired')
	project.append('categoryID')
	project.append('company')
	project.append('companyID')
	project.append('condition')
	project.append('conditionType')
	project.append('cpi')
	project.append('csi')
	project.append('spi')
	project.append('description')
	project.append('durationMinutes')
	project.append('eac')
	project.append('enteredByID')
	project.append('entryDate')
	project.append('estCompletionDate')
	project.append('estStartDate')
	project.append('groupID')
	project.append('ID')
	project.append('lastUpdateDate')
	project.append('name')
	project.append('owner')
	project.append('ownerID')
	project.append('percentComplete')
	project.append('plannedCompletionDate')
	project.append('plannedLaborCost')
	project.append('plannedStartDate')
	project.append('portfolioID')
	project.append('priority')
	project.append('projectedStartDate')
	project.append('projectedCompletionDate')
	project.append('risk')
	project.append('referenceNumber')
	project.append('sponsorID')
	project.append('status')
	project.append('statusUpdate')
	project.append('template')
	project.append('URL')
	project.append('workRequired')
	project.append('portfolio')
	project.append('DE:Account Name')
	project.append('DE:Additional Client Notes')
	project.append('DE:Backup Voice Talent')
	project.append('DE:Benefits Eligible Employees')
	project.append('DE:Brand Management Designer/Art Director')
	project.append('DE:Brand Management Writer/Client Content Strategist')
	project.append('DE:Calculated Completion Date')
	project.append('DE:Customer Goals/Objectives')
	project.append('DE:Customer Hard Deadline')
	project.append('DE:Current Velocity')
	project.append('DE:Date Logged')
	project.append('DE:Days In Queue')
	project.append('DE:Days In Queue Color')
	project.append('DE:Days with Voice Actor 1')
	project.append('DE:Days with Voice Actor 2')
	project.append('DE:Days with Voice Actor 3')
	project.append('DE:Days with Voice Actor 4')
	project.append('DE:Days with Voice Actor 5')
	project.append('DE:Days with Voice Actor 6')
	project.append('DE:Days with Voice Actor 7')
	project.append('DE:Days with Voice Actor Final Cut')
	project.append('DE:Days with Voice Actor Initial Development')
	project.append('DE:Driver for Hard Deadline')
	project.append('DE:Driver for Soft Deadline')
	project.append('DE:Effective Date')
	project.append('DE:G1/G2')
	project.append('DE:Line Item ID')
	project.append('DE:List demos shown and customer feedback')
	project.append('DE:New Customer/Upsell/Renewal')
	project.append('DE:New Titles Requested')
	project.append('DE:Opportunity ID')
	project.append('DE:Order Number')
	project.append('DE:Persona')
	project.append('DE:Product Code')
	project.append('DE:Product Group')
	project.append('DE:OPERATIONAL | Product Line')
	project.append('DE:Product Line')
	project.append('DE:Product Title')
	project.append('DE:Product Video Category')
	project.append('DE:Product Video Subcategory')
	project.append('DE:Production Tool')
	project.append('DE:Program Start Date')
	project.append('DE:Project Total Velocity')
	project.append('DE:Renewal Date')
	project.append('DE:Requested Kickoff Call Timing')
	project.append('DE:Role Suffix')
	project.append('DE:Script Word Count')
	project.append('DE:Soft Deadline')
	project.append('DE:Substatus')
	project.append('DE:Target Deadline and Drivers')
	project.append('DE:Target Language')
	project.append('DE:Template Type')
	project.append('DE:Template Version')
	project.append('DE:To Translation Services?')
	project.append('DE:Voice In 1')
	project.append('DE:Voice In 2')
	project.append('DE:Voice In 3')
	project.append('DE:Voice In 4')
	project.append('DE:Voice In 5')
	project.append('DE:Voice In 6')
	project.append('DE:Voice In 7')
	project.append('DE:Voice in Final Cut')
	project.append('DE:Voice In Initial Development')
	project.append('DE:Voice Out 1')
	project.append('DE:Voice Out 2')
	project.append('DE:Voice Out 3')
	project.append('DE:Voice Out 4')
	project.append('DE:Voice Out 5')
	project.append('DE:Voice Out 6')
	project.append('DE:Voice Out 7')
	project.append('DE:Voice Out Final Cut')
	project.append('DE:Voice Out Initial Development')
	project.append('DE:Voice Talent')
	project.append('DE:Video Type')
	project.append('DE:Video Points')

	return(project)

def update_fields():

	update=[]
	update.append('ID')
#	update.append('allowedActions')
#	update.append('enteredByID')
#	update.append('enteredByName')
#	update.append('entryDate')
#	update.append('iconName')
#	update.append('iconPath')
#	update.append('message')
#	update.append('refName')
#	update.append('refObjCode')
#	update.append('refObjID')
#	update.append('styledMessage')
#	update.append('subMessage')
#	update.append('subObjCode')
#	update.append('subObjID')
#	update.append('threadID')
#	update.append('topName')
#	update.append('topObjCode')
#	update.append('topObjID')
#	update.append('updateActions')
#	update.append('updateObjCode')
#	update.append('updateObjID')
#	update.append('updateType')

	return(update)

def template_fields():

	template=[]
	template.append('ID')
	template.append('name')

	return(template)

def template_task_fields():

	template=[]
	template.append('templateID')
	template.append('name')

	return(template)

def task_fields():

	task=[]
	task.append('actualCompletionDate')
	task.append('actualDuration')
	task.append('actualDurationMinutes')
	task.append('actualCost')
	task.append('actualExpenseCost')
	task.append('actualLaborCost')
	task.append('actualStartDate')
	task.append('actualWork')
	task.append('actualWorkRequired')
	task.append('approvalEstStartDate')
	task.append('approvalPlannedStartDate')
	task.append('approvalProjectedStartDate')
	task.append('assignedToID')
	task.append('commitDate')
	task.append('completionPendingDate')
	task.append('costAmount')
	task.append('costType')
	task.append('cpi')
	task.append('csi')
	task.append('duration')
	task.append('durationMinutes')
	task.append('durationType')
	task.append('durationUnit')
	task.append('entryDate')
	task.append('estCompletionDate')
	task.append('estStartDate')
	task.append('handoffDate')
	task.append('ID')
	task.append('lastUpdateDate')
	task.append('name')
	task.append('numberOfChildren')
	task.append('originalDuration')
	task.append('percentComplete')
	task.append('plannedCost')
	task.append('plannedCompletionDate')
	task.append('plannedDuration')
	task.append('plannedDurationMinutes')
	task.append('plannedExpenseCost')
	task.append('plannedLaborCost')
	task.append('plannedStartDate')
	task.append('priority')
	task.append('progressStatus')
	task.append('projectedCompletionDate')
	task.append('projectedDurationMinutes')
	task.append('projectedStartDate')
	task.append('projectID')
	task.append('remainingDurationMinutes')
	task.append('roleID')
	task.append('spi')
	task.append('status')
	task.append('taskNumber')
	task.append('teamID')
	task.append('wbs')
	task.append('work')
	task.append('workRequired')
	task.append('workRequiredExpression')

	return(task)

def role_fields():

	role=[]
	role.append('ID')
	#role.append('entryDate')
	role.append('name')

	return(role)

def user_fields():

	user=[]
	user.append('ID')
	user.append('name')
	user.append('roleID')
	user.append('reservedTimes')

	return(user)

def user_act_fields():

	user_act=[]
	user_act.append('ID')
	#user_act.append('customerID')
	#user_act.append('name')
	#user_act.append('entryDate')
	#user_act.append('lastUpdateDate')
	#user_act.append('userID')
	user_act.append('value')

	return(user_act)

def reserved_times_fields():

	reserved_times=[]
	reserved_times.append('ID')
	reserved_times.append('endDate')
	reserved_times.append('startDate')
	reserved_times.append('taskID')
	reserved_times.append('userID')

	return(reserved_times)

def note_fields():

	note=[]
	note.append('ID')
	note.append('attachObjCode')
	note.append('attachObjID')
	note.append('auditText')
	note.append('auditType')
	note.append('entryDate')
	note.append('customerID')
	note.append('ownerID')
	note.append('extRefID')
	note.append('parentJournalEntryID')
	note.append('parentNoteID')
	note.append('numReplies')
	note.append('isMessage')
	note.append('noteText')
	note.append('numReplies')
	note.append('subject')
	note.append('threadID')
	note.append('threadDate')
	note.append('objID')
	note.append('topNoteObjCode')
	note.append('topObjID')

	return(note)

def hour_fields():

	hour=[]
	hour.append('ID')
	hour.append('actualCost')
	hour.append('approvedByID')
	hour.append('approvedOnDate')
	hour.append('entryDate')
	hour.append('hours')
	hour.append('hourTypeID')
	hour.append('ownerID')
	hour.append('roleID')
	hour.append('projectID')
	hour.append('taskID')
	hour.append('status')
	hour.append('opTaskID')

	return(hour)

def hour_type_fields():

	hour_type=[]
	hour_type.append('ID')
	hour_type.append('description')
	hour_type.append('extRefID')
	hour_type.append('isActive')
	hour_type.append('name')
	hour_type.append('nameKey')
	hour_type.append('objID')

	return(hour_type)

def timesheet_fields():

	timesheet=[]
	timesheet.append('ID')
	timesheet.append('endDate')
	timesheet.append('extRefID')
	timesheet.append('hasNotes')
	timesheet.append('lastUpdateDate')
	timesheet.append('overtimeHours')
	timesheet.append('regularHours')
	timesheet.append('startDate')
	timesheet.append('status')
	timesheet.append('timesheetProfileID')
	timesheet.append('totalHours')
	timesheet.append('userID')
#	timesheet.append('approverCost')
#	timesheet.append('availableActions')
#	timesheet.append('customerID')
#	timesheet.append('displayName')
#	timesheet.append('hoursDuration')
#	timesheet.append('lastNoteID')
#	timesheet.append('lastUpdateByID')

	return(timesheet)

def non_work_day_fields():

	non_work_day=[]
	non_work_day.append('ID')
	#non_work_day.append('nonWorkDay')
	#non_work_day.append('objID')
	#non_work_day.append('Scope')
	#non_work_day.append('scheduleDay')
	#non_work_day.append('scheduleID')
	#non_work_day.append('userID')

	return(non_work_day)

def task_ranking(client,cur_datetime,update_rankings,timing_bin,company_df,attask_input_df):

	#############################################
	# Created by DKJ 10/10/2014
	#
	# Must run after 'task_buffer_PROD.py
	# Will not run by itself
	#############################################
	
	bin_correction_multiplier = 1
	
	##############################
	# Should we update rankings
	##############################
	if 'update_rankings' in locals():
		LOG.info("'update_rankings' variable exists:\n\n")
	else:
		LOG.info("TERMINATE PROGRAM ... 'update_rankings' variable DOES NOT exist\n\n")
		sys.exit()
	
	#######################
	# PARENT TASK LOOKUP
	#######################
	parent_task_lookup = []
	parent_task_lookup.append('INITIAL DEVELOPMENT')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #1')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #2')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #3')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #4')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #5')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #6')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #7')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #8')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #9')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #10')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #11')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #12')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #13')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #14')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #15')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #16')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #17')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #18')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #19')
	parent_task_lookup.append('VIDEO REVIEW CYCLE #20')
	parent_task_lookup.append('ADDITIONAL VIDEO EDITS')
	parent_task_lookup.append('FINAL APPROVAL & PUBLISH')
	
	#######################################################################################################
	# Rank the parent tasks by priority
	#######################################################################################################
	parent_task_rank = [-1] * len(attask_input_df) 
	for i in range(0,len(attask_input_df)):
		try:
			parent_task_rank[i] = parent_task_lookup.index(attask_input_df.PARENT_task[i]) + 1
		except:
			parent_task_rank[i] = len(parent_task_lookup) + 1
			
	#for i in range(0,len(parent_task_lookup)):
	#	Nidx = all_indices(parent_task_lookup[i],attask_input_df.PARENT_task)
	#
	#	for j in range(0,len(Nidx)):
	#		parent_task_rank[Nidx[j]] = i
	
	attask_input_df = attask_input_df.join(pd.DataFrame(parent_task_rank,columns=['PARENT_task_rank']))
	
	#####################################################################################
	# Create date field which is hard_deadline .. if None, it becomes soft_deadline
	#####################################################################################
	best_close_date_type = []
	best_close_date = []
	day_delta_best = []
	for i in range(0,len(attask_input_df)):
		if (attask_input_df['hard_deadline'][i] != None and pd.notnull(attask_input_df['hard_deadline'][i])):
			best_close_date.append(attask_input_df['hard_deadline'][i])
			best_close_date_type.append('hard')
			day_delta_best.append(attask_input_df['day_delta_hard_deadline'][i])
			if ('NVIDIA' in attask_input_df.project_name[i]):
				LOG.info(':{:>4}:{}:{}\n'.format(i,attask_input_df.project_name[i],attask_input_df['hard_deadline'][i]))
				#sys.exit()
		else: 
			if (attask_input_df['soft_deadline'][i] != None and pd.notnull(attask_input_df['soft_deadline'][i])):
				best_close_date.append(attask_input_df['soft_deadline'][i])
				best_close_date_type.append('soft')
				day_delta_best.append(attask_input_df['day_delta_soft_deadline'][i])
			else:
				best_close_date.append(None)
				best_close_date_type.append(None)
				day_delta_best.append(None)
	
	attask_input_df = attask_input_df.join(pd.DataFrame(best_close_date,columns=['best_close_date']))
	attask_input_df = attask_input_df.join(pd.DataFrame(best_close_date_type,columns=['best_close_date_type']))
	attask_input_df = attask_input_df.join(pd.DataFrame(day_delta_best,columns=['day_delta_best']))
	
	################################################
	# Add total number of open projects per company
	################################################
	#company_group = company_df.groupby(['company']).aggregate(np.sum).reset_index()
	company_group = pd.DataFrame(attask_input_df[['ID','company']]).groupby(['company']).count().reset_index().rename(columns={"ID":"Nlibrary"})
	attask_input_df = pd.merge(attask_input_df,company_group,how='left',right_on='company',left_on='company').reset_index(drop=True)
	
	##################################################
	# Bin sub-groups for each customer with >4 videos
	# KLUGE: best to filter attask_input_df ... but giving 
	# issues 
	##################################################
	dict_completion = {}
	bin_correction = [0] * len(attask_input_df)
	for i in range(0,len(company_group)):
		Pidx = []
		TMP_completion_date = []
		TMP_best_close_date = []
		Cidx = all_indices(company_group['company'][i],attask_input_df['company'])	
		for j in range(0,len(Cidx)):
			TMP_completion_date.append(attask_input_df['PARENT_completion_date_planned'][Cidx[j]])			
			TMP_best_close_date.append(attask_input_df['best_close_date'][Cidx[j]])			
	
		dict_completion['Cidx'] = Cidx
		dict_completion['PARENT_completion_date'] = TMP_completion_date 
		dict_completion['best_close_date'] = TMP_best_close_date 
		completion_df = pd.DataFrame(dict_completion).sort(['best_close_date','PARENT_completion_date'],ascending=[1,1]).reset_index(drop=True)	
	
		for j in range(0,len(completion_df)):
			idx = completion_df.ix[j]['Cidx']
			if (attask_input_df['Nlibrary'][idx] != 'nan'):
				#bin_correction[idx] = attask_input_df['Nlibrary'][idx] - int(j/float(bin_correction_multiplier))*bin_correction_multiplier 
				bin_correction[idx] = attask_input_df['Nlibrary'][idx] - int(j/float(bin_correction_multiplier))*bin_correction_multiplier 
	
	attask_input_df = attask_input_df.join(pd.DataFrame(bin_correction,columns=['bin_correction']))
	
	#############################
	# Find indices for each bin 
	#############################
	data = np.array(attask_input_df['day_delta_best'])
	bins = np.linspace(-500,500,1000/timing_bin+1)
	bin_idx = np.digitize(data,bins)
	
	##### Set null values to -1
	isnull = list(attask_input_df[pd.isnull(attask_input_df['day_delta_best'])].index)
	bin_idx[isnull] = -1
	
	bin_idx_BAK = list(bin_idx)
	
	########################################################################################################################
	# Add correction to bin_idx
	# a) 1 week per # of videos open
	# b) 1st video does not count
	#
	# How to split large libraries
	# c) Split entire library into groups of 4 based on current completion date ... groups defined by 'bin_correction_multiplier'
	#	1st set is 1 video 	... 1 video with the earliest completion dates (1)
	#	2nd set of 5 videos ... 5 videos with the next earliest completion dates (1-6)
	#	3rd set of 5 videos ... 5 videos with the next earliest completion dates (7-11)
	#	etc ....
	#
	# d) Distribute groups across timeframe
	#
	#	Date Configuration for 16 video library	
	#	--------------------------------------------------------------------------------------------------------------------
	#	|1st set 				|2nd set		 		|3rd set				|4th set				| Hard Deadline			
	#	|Hard Deadline - 16 wks	|Hard Deadline - 12 wks |Hard Deadline - 8 wks	|Hard Deadline - 4 wks
	#
	########################################################################################################################
	for i in range(0,len(bin_idx)):
		if (bin_idx[i] != -1):
			#bin_idx[i] = bin_idx[i] + attask_input_df['Nlibrary'][i] - 1
			bin_idx[i] = bin_idx[i] + attask_input_df['bin_correction'][i] - 1
	
	attask_input_df = attask_input_df.join(pd.DataFrame(bin_idx,columns=['BIN_day_delta_best']))
	
	###################################################################################################
	# Rank output
	# 1) BIN_day_delta_best (determined by timing_bins
	# 2) best_close_date_type (hard vs soft) 
	# 3) PARENT_task 
	# 4) day_delta ... (>0 to <0 ... <0 means that you are upstream of the deadline)
	#
	# If timing_bin = 10, may want to add 'day_delta_best' 
	###################################################################################################
	#attask_input_df = attask_input_df.sort(['BIN_day_delta_best','best_close_date_type','day_delta_best','ranking_flag','day_delta_w_remaining_buffer'],ascending=[0,1,0,0,0]).reset_index(drop=True)
	attask_input_df = attask_input_df.sort(['BIN_day_delta_best','best_close_date_type','PARENT_task_rank','day_delta_best'],ascending=[0,1,1,0]).reset_index(drop=True)
	
	#if (timing_bin == 5):
	#print(attask_input_df[['project_status','BIN_day_delta_best','best_close_date_type','ranking_flag','day_delta_w_remaining_buffer','day_delta_w_buffer']])
	
	if update_rankings:
		attask_update_success = []
		for i in range(0,len(attask_input_df)):	
			#############################################
			# Update 'Project Ranking' & 'Ranking Flag'
			#############################################
			Nrank = i + 1
			LOG.info('{0:>5} of {1:>5} ... {2:>4} ...  {3:>4} .  {4:>4} .  {5:>6} .  {6:>6} ...  {7:>100}'.format(i+1,len(attask_input_df),Nrank, \
											attask_input_df.BIN_day_delta_best[i], \
											attask_input_df.best_close_date_type[i], \
											attask_input_df.day_delta_best[i], \
											attask_input_df.status[i], \
											attask_input_df.project_name[i].encode('ascii','ignore') ) )
			cur_project = attask_input_df.project_name[i]
	
			try:
				cur_project = AtTaskObject(client.put(ObjCode.PUT_PROJECT,attask_input_df.ID[i],{'DE:Project Ranking':Nrank}),client)
				LOG.info(' ... Project Ranking/Ranking Flag UPDATED')
				attask_update_success.append(1)
			except:
				LOG.info(' ... Project Ranking/Ranking Flag Update Failed')
				attask_update_success.append(0)
	
		####################################
		# Add attask_update_success column
		####################################			
		attask_input_df = attask_input_df.join(pd.DataFrame(attask_update_success,columns=['attask_update_success']))
	
		###############
		# Output File
		###############
		attask_input_df.to_csv('./output/attask_input_df_' + cur_datetime.strftime('%Y%m%d') +'.csv',encoding='utf-8')
	
	###############################
	# Create xlsx file
	###############################
	wb = Workbook()
	ws = wb.create_sheet(0)
	wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
	ws.title = 'project_ranking'
	
	####################
	# Format xlsx file
	####################
	ws.cell(row = 0, column = 0).style.alignment.wrap_text = True
	ws.merge_cells('A1:C1')
	ws.merge_cells('D1:E1')
	for i in range(0,14):
		ws.cell(row = 1, column = i).style.fill.fill_type = Fill.FILL_SOLID 
		ws.cell(row = 1, column = i).style.fill.start_color.index = Color.DARKBLUE 
		ws.cell(row = 1, column = i).style.font.color = Color(Color.WHITE) 
	for i in range(0,14):
		ws.cell(row = 0, column = 0).style.alignment.horizontal = "center" 
		ws.cell(row = 0, column = 0).style.alignment.vertical = "center" 
	
	#for j in range(3,30):
	#	ws.cell(row = j, column = 5).style.number_format.format_code = "$#,##0" 
		
	ws.cell(row = 0,column = 0).value = 'Report Creation Time'
	ws.cell(row = 0,column = 3).value = cur_datetime
	ws.cell(row = 0,column = 7).value = '=SUBTOTAL(3,H3:H1000000)'
	ws.cell(row = 1,column = 0).value = 'Rank'
	ws.cell(row = 1,column = 1).value = 'Company'
	ws.cell(row = 1,column = 2).value = 'Library Size'
	ws.cell(row = 1,column = 3).value = 'Completion Date'
	ws.cell(row = 1,column = 4).value = 'Project Name'
	ws.cell(row = 1,column = 5).value = 'Project Status'
	ws.cell(row = 1,column = 6).value = 'CSM'
	ws.cell(row = 1,column = 7).value = 'Current Task'
	ws.cell(row = 1,column = 8).value = 'Parent Task'
	ws.cell(row = 1,column = 9).value = 'Hard Deadline'
	ws.cell(row = 1,column = 10).value = 'Soft Deadline'
	ws.cell(row = 1,column = 11).value = 'Day_delta_best'
	ws.cell(row = 1,column = 12).value = 'BIN_day_delta_best'
	ws.cell(row = 1,column = 13).value = 'Best_close_date_type'
	
	########################
	# Output attask_input_df
	########################
	Icsm = 0
	for i in range(0,len(attask_input_df)):
		ws.cell(row = (Icsm+2), column = 0).value = i+1
		ws.cell(row = (Icsm+2), column = 1).value = attask_input_df.company[i] 
		ws.cell(row = (Icsm+2), column = 2).value = attask_input_df.Nlibrary[i] 
		ws.cell(row = (Icsm+2), column = 3).value = str(attask_input_df.PARENT_completion_date_planned[i])[0:10]
		try:
			ws.cell(row = (Icsm+2), column = 4).value = attask_input_df.project_name[i].encode('ascii','ignore') 
		except:
			LOG.info("Skip");
		ws.cell(row = (Icsm+2), column = 5).value = attask_input_df.status[i] 
		ws.cell(row = (Icsm+2), column = 6).value = attask_input_df.CSM[i] 
		ws.cell(row = (Icsm+2), column = 7).value = attask_input_df.current_task[i] 
		
		idx = 7   
		idx = idx + 1	
		try:
			if (pd.notnull(attask_input_df.PARENT_task[i])):
				ws.cell(row = (Icsm+2), column = idx).value = attask_input_df.PARENT_task[i] 
			else:
				ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
		except:
			ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
	
		idx = idx + 1	
		try:
			if (pd.notnull(attask_input_df.hard_deadline[i])):
				ws.cell(row = (Icsm+2), column = idx).value = attask_input_df.hard_deadline[i] 
			else:
				ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
		except:
			ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
	
		idx = idx + 1
		try:
			if (pd.notnull(attask_input_df.soft_deadline[i])):
				ws.cell(row = (Icsm+2), column = idx).value = attask_input_df.soft_deadline[i] 
			else:
				ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
		except:
			ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
	
		idx = idx + 1
		try:
			if (pd.notnull(attask_input_df.day_delta_best[i])):
				ws.cell(row = (Icsm+2), column = idx).value = attask_input_df.day_delta_best[i] 
			else:
				ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
		except:
			ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
	
		idx = idx + 1
		ws.cell(row = (Icsm+2), column = idx).value = attask_input_df.BIN_day_delta_best[i] 
	
		idx = idx + 1
		try:
			if (pd.notnull(attask_input_df.best_close_date_type[i])):
				ws.cell(row = (Icsm+2), column = idx).value = attask_input_df.best_close_date_type[i] 
			else:
				ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
		except:
			ws.cell(row = (Icsm+2), column = idx).value = 'NA' 
	
		Icsm = Icsm + 1
	
	Nrow = ws.get_highest_row();
	LOG.info('Nrow = {}'.format(Nrow))
	ws.auto_filter = "A2:N" + str(Nrow) # Turn on Autofilter 
	wb.save('./output/project_ranking_' + cur_datetime.strftime('%Y%m%d') +'.xlsx')
		
