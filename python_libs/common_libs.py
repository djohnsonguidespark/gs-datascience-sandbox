#! /usr/bin/env python

import sys
import re
import json
import csv
import collections
from openpyxl import load_workbook
from openpyxl.reader.excel import Workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
from datetime import datetime, timedelta
from subprocess import call
import MySQLdb as mdb
import pandas as pd
import numpy as np 
import math 

# Logging
import log_libs as log
LOG = log.init_logging()

def printf(format_str, *args):
    sys.stdout.write(format_str % args)

def printf_NEW(fi,format_str, *args):
	format_out = format_str % args
	try:
		filename = fi.f_code.co_filename
		try:
			file_out = filename.split("/")[len(filename.split("/"))-1]
		except:
			file_out = fileinfo.filename
		if (fi.f_code.co_name == '<module>'):
			line_out = "[%s][Line = %5s]: %s" % (file_out,fi.f_lineno,format_out)
		else:
			line_out = "[%s][func: %s][Line = %5s]: %s" % (file_out,fi.f_code.co_name,fi.f_lineno,format_out)
	except:
		line_out = format_out
	sys.stdout.write(line_out)

def print_full(x):
    pd.set_option('display.max_rows', len(x))
    print(x)
    pd.reset_option('display.max_rows')

def createDF_from_MYSQL_query(query_in):

	con = None
	con = mdb.connect('localhost','root','','');
	cur = con.cursor()

	cur.execute(query_in)
	tmp_data = cur.fetchall()

	tmp_dict = {}
	for i in range(0,len(cur.description)):
		tmp_list = []
		for j in range(0,len(tmp_data)):
			tmp_list.append(tmp_data[j][i])

		tmp_dict[cur.description[i][0]] = tmp_list
 
	return(pd.DataFrame(tmp_dict))

def calculate_BENCHMARK(name,bin_out_filter,data_in_df):

	unique_bin = [x for x in sorted(list(set(bin_out_filter))) if x is not None]

	#unique_industry.pop(0)  ## Remove None	j

	bin_list = []
	days_since_OE = []
	mean_data_out = []
	median_data_out = []
	for i in range(0,len(unique_bin)):
		idx_bin = all_indices(unique_bin[i],bin_out_filter)
		cur_bin_out = data_in_df[data_in_df.columns[idx_bin]]

		for j in range(0,len(cur_bin_out)):
			bin_list.append(unique_bin[i])
			days_since_OE.append(j)
			mean_data_out.append(np.mean(cur_bin_out.ix[j]))	
			median_data_out.append(np.median(cur_bin_out.ix[j]))	

	bin_reach = {}
	bin_reach[name] = bin_list 
	bin_reach['days_since_OE'] = days_since_OE
	bin_reach['median_reach'] = median_data_out
	#bin_reach['mean_reach'] = mean_data_out 
	bin_reach_df = pd.DataFrame(bin_reach)
	return bin_reach_df 


def create_BEE_bin(in_df):
	BEE_bin = [0] * len(in_df)
	for i in range(0,len(BEE_bin)):
		if (in_df['BEEs'][i] > 0 and in_df['BEEs'][i] <= 999):
			BEE_bin[i] = '0-999'
		elif (in_df['BEEs'][i] >= 1000 and in_df['BEEs'][i] <= 4999):
			BEE_bin[i] = '1000-4999'
		elif (in_df['BEEs'][i] >= 5000 and in_df['BEEs'][i] <= 49999):
			BEE_bin[i] = '5000-49999'
		elif (in_df['BEEs'][i] >= 50000):
			BEE_bin[i] = '50000+'

	in_df = in_df.join(pd.DataFrame(BEE_bin))
	in_df = in_df.rename(columns={0:'BEE_bin'})

	return in_df

def create_Library_bin(in_df):
	library_bin = [0] * len(in_df)
	for i in range(0,len(library_bin)):
		if (in_df['Nvideo'][i] == 1):
			library_bin[i] = '01'
		elif (in_df['Nvideo'][i] >= 2 and in_df['Nvideo'][i] <= 3):
			library_bin[i] = '02-03'
		elif (in_df['Nvideo'][i] >= 4 and in_df['Nvideo'][i] <= 9):
			library_bin[i] = '04-09'
		elif (in_df['Nvideo'][i] >= 10 and in_df['Nvideo'][i] <= 19):
			library_bin[i] = '10-19'
		elif (in_df['Nvideo'][i] >= 20):
			library_bin[i] = '20+'

	in_df = in_df.join(pd.DataFrame(library_bin))
	in_df = in_df.rename(columns={0:'Nvideo_bin'})

	return in_df

def createDF_from_MYSQL(DBNAME,TABLENAME,query_in=None):

	con = None
	con = mdb.connect('localhost','root','','');
	cur = con.cursor()

	query_col = "SELECT `COLUMN_NAME` \
				 FROM `INFORMATION_SCHEMA`.`COLUMNS` \
				 WHERE `TABLE_SCHEMA`='%s' \
				 AND `TABLE_NAME`='%s'" % (DBNAME,TABLENAME)
	cur.execute(query_col)
	tmp_data_col = cur.fetchall()

	if (query_in is not None):
		query = query_in % (DBNAME,TABLENAME)
	else:
		query = "SELECT * FROM %s.%s" % (DBNAME,TABLENAME)

	cur.execute(query)
	tmp_data = cur.fetchall()

	tmp_dict = {}
	for i in range(0,len(tmp_data_col)):
		tmp_list = []
		for j in range(0,len(tmp_data)):
			tmp_list.append(tmp_data[j][i])

		tmp_dict[tmp_data_col[i][0]] = tmp_list
 
	return(pd.DataFrame(tmp_dict))

def createDF_from_CSV(csv_file):

	f = open(csv_file,'rU')
	data_in = []
	try:
		reader = csv.DictReader(f)
		for row in reader:
			data_in.append(row)
	finally:
		f.close()

	keys = list(sorted(list(row.keys()))) 

	data_out_dict = {}
	for i in range(0,len(keys)):
		data_out_dict[keys[i]] = map((lambda item: item[keys[i]]),data_in)		

	return(data_out_dict)

def detectMobile(user_agent):
	reg_b = re.compile(r"(android|bb\\d+|meego).+mobile|avantgo|bada\\/|blackberry|blazer|compal|elaine|fennec|hiptop|iemobile|ip(hone|od)|iris|kindle|lge |maemo|midp|mmp|mobile.+firefox|netfront|opera m(ob|in)i|palm( os)?|phone|p(ixi|re)\\/|plucker|pocket|psp|series(4|6)0|symbian|treo|up\\.(browser|link)|vodafone|wap|windows (ce|phone)|xda|xiino", re.I|re.M)
	reg_v = re.compile(r"1207|6310|6590|3gso|4thp|50[1-6]i|770s|802s|a wa|abac|ac(er|oo|s\\-)|ai(ko|rn)|al(av|ca|co)|amoi|an(ex|ny|yw)|aptu|ar(ch|go)|as(te|us)|attw|au(di|\\-m|r |s )|avan|be(ck|ll|nq)|bi(lb|rd)|bl(ac|az)|br(e|v)w|bumb|bw\\-(n|u)|c55\\/|capi|ccwa|cdm\\-|cell|chtm|cldc|cmd\\-|co(mp|nd)|craw|da(it|ll|ng)|dbte|dc\\-s|devi|dica|dmob|do(c|p)o|ds(12|\\-d)|el(49|ai)|em(l2|ul)|er(ic|k0)|esl8|ez([4-7]0|os|wa|ze)|fetc|fly(\\-|_)|g1 u|g560|gene|gf\\-5|g\\-mo|go(\\.w|od)|gr(ad|un)|haie|hcit|hd\\-(m|p|t)|hei\\-|hi(pt|ta)|hp( i|ip)|hs\\-c|ht(c(\\-| |_|a|g|p|s|t)|tp)|hu(aw|tc)|i\\-(20|go|ma)|i230|iac( |\\-|\\/)|ibro|idea|ig01|ikom|im1k|inno|ipaq|iris|ja(t|v)a|jbro|jemu|jigs|kddi|keji|kgt( |\\/)|klon|kpt |kwc\\-|kyo(c|k)|le(no|xi)|lg( g|\\/(k|l|u)|50|54|\\-[a-w])|libw|lynx|m1\\-w|m3ga|m50\\/|ma(te|ui|xo)|mc(01|21|ca)|m\\-cr|me(rc|ri)|mi(o8|oa|ts)|mmef|mo(01|02|bi|de|do|t(\\-| |o|v)|zz)|mt(50|p1|v )|mwbp|mywa|n10[0-2]|n20[2-3]|n30(0|2)|n50(0|2|5)|n7(0(0|1)|10)|ne((c|m)\\-|on|tf|wf|wg|wt)|nok(6|i)|nzph|o2im|op(ti|wv)|oran|owg1|p800|pan(a|d|t)|pdxg|pg(13|\\-([1-8]|c))|phil|pire|pl(ay|uc)|pn\\-2|po(ck|rt|se)|prox|psio|pt\\-g|qa\\-a|qc(07|12|21|32|60|\\-[2-7]|i\\-)|qtek|r380|r600|raks|rim9|ro(ve|zo)|s55\\/|sa(ge|ma|mm|ms|ny|va)|sc(01|h\\-|oo|p\\-)|sdk\\/|se(c(\\-|0|1)|47|mc|nd|ri)|sgh\\-|shar|sie(\\-|m)|sk\\-0|sl(45|id)|sm(al|ar|b3|it|t5)|so(ft|ny)|sp(01|h\\-|v\\-|v )|sy(01|mb)|t2(18|50)|t6(00|10|18)|ta(gt|lk)|tcl\\-|tdg\\-|tel(i|m)|tim\\-|t\\-mo|to(pl|sh)|ts(70|m\\-|m3|m5)|tx\\-9|up(\\.b|g1|si)|utst|v400|v750|veri|vi(rg|te)|vk(40|5[0-3]|\\-v)|vm40|voda|vulc|vx(52|53|60|61|70|80|81|83|85|98)|w3c(\\-| )|webc|whit|wi(g |nc|nw)|wmlb|wonu|x700|yas\\-|your|zeto|zte\\-", re.I|re.M)

	mobile=0	
	if (user_agent != None and user_agent != ''):
		if ('GUIDESPARK' not in user_agent.upper()):
			b = reg_b.search(user_agent)
			v = reg_v.search(user_agent[0:4])
			if b or v:
				mobile=1	
				LOG.debug(":%s:",user_agent)
		else:
			if ('MOBILE' in user_agent.upper()):
				mobile=1
				LOG.debug(":%s:",user_agent)

	## mobile=1 .. MOBILE 
	## mobile=0 .. Desktop
	return(mobile)

def createPenetration(library,bee,video_per_session,input_views):

	pen = {}
	pen['bee'] = int(bee[0])
	pen['library'] = library
	pen['video_per_session'] = video_per_session 
	views = []
	pen['views'] = input_views 

	return(pen)

def concat_list(list1,list2,separator):
	output = []
	for i in range(0,len(list1)):
		LOG.debug(":%s:%s:%s:",i,list1[i],list2[i])
		try:
			output.append(list1[i].encode('latin1').replace('\x92','') + separator + list2[i]) 
		except:
			LOG.debug("CONCAT:%s:%s:%s:",i,list1[i],list2[i])
			output.append(list1[i].encode('latin1').replace('\x92','') + separator) 
	return output

def check_value(str_in,attribute):

	try:
		if (attribute != 'mkto_si__Sales_Insight__c'): 
			output = str_in[attribute]
		else:
			output = str_in[attribute].split('?id=')[1].split('"')[0] 
	except:
		output = "" 

	return output

def check_if_nan(str_in):
	try:
		if (pd.isnull(str_in) == True):
			output = None
		else:
			output = str_in 
	except:
		output = None

	return output

def replace_unicode(str_in):
	try:
		output = str_in.encode('ascii','unicode').replace("'","").replace("\n","")
	except:
		output = None

	return output

def ListElement0(val_in):
	try:
		if (type(val_in) is list):
			output = val_in[0]
		else:
			output = val_in 
	except:
		output = None		

	return output

def check_NONE(str_in,column):
	if (str_in != None):
		output = str_in[column] 
	else:
		output = None

	return output

#########################################################################
# print_list_substring -> takes in 
# 1) substring -> substring to determine indices to print
# 2) str_list_substring -> str list that determines indices to print
# 3) str_to_print -> str list that you want printed with the index list
#########################################################################
def print_list_substring(substring,str_list_substring,str_to_print):
	if (str_to_print == None):
		str_to_print = str_list_substring

	index_list = all_substring(substring,str_list_substring)
	for i in range(0,len(index_list)):
	  	LOG.info(":%s:",str_to_print[index_list[i]])

def print_list(print_list,str_to_print):
	for i in range(0,len(print_list)):
	  	LOG.info(":%s:",str_to_print[print_list[i]])
	LOG.info("Total Elements :%s:",len(print_list) )

def all_indices(string,str_list):
  	LOG.debug(":%s:%s:",type(string),string)

	if isinstance(string,str):
  		LOG.debug("INSIDE:%s:",string.upper())
		return ([i for i, x in enumerate(str_list) if x is not None and x.upper() == string.upper()]) 
	else:
		return ([i for i, x in enumerate(str_list) if x is not None and x == string]) 

def all_indices_CASE_SENSITIVE(string,str_list):
	return ([i for i, x in enumerate(str_list) if x == string]) 

#def all_indicesFLOAT(input,float_list):
#  	LOG.info(":%s:%s",strings,substring)
#	return ([i for i, x in enumerate(str_list) if x == string]) 

def all_substring(substring, str_list):
  	#LOG.debug(":%s:%s",strings,substring)

	#### Replace NONE with 'ZZZZZZZ'
	idx = all_indices(None,str_list)
	for i in range(0,len(idx)):
		str_list[idx[i]] = 'ZZZZZZ'	
	return ([i for i, x in enumerate(str_list) if substring in x])

def intersect(a,b):
	return list(set(a) & set(b))

def union(a,b):
	return list(set(a) | set(b))

def extra_val(a,b):
	return list(set(a).difference(b))

def list_summary(input_list):

	unique_list = list(set(input_list))

	browser_name = []
	browser_count = []
	for i in range(0,len(unique_list)):
		browser_name.append(unique_list[i])
		browser_count.append(input_list.count(unique_list[i]) )

	dict_output = {}
	dict_output['browser_name'] = browser_name 
	dict_output['browser_count'] = browser_count
	df = pd.DataFrame(dict_output)
	return(df.sort(['browser_count'],ascending=[0]) )

def id_unique_symbols(title,sheet,input_val):
	unique_symbols = []
	for j in range(0,len(input_val)):
		if (input_val[j].game_title == title and input_val[j].sheet == sheet):
			LOG.debug(":%s:",input_val[j].win_symbol)
			for win in input_val[j].win_symbol.split(";"):
				for symbol in win.split(","):
					if (symbol != ''):
						unique_symbols.append(symbol)

	return sorted(list(set(unique_symbols)))
	
def create_symbol_string(search_str,input_val):		

	Tsymbol = ""
	#Tint = first_substring(input_val,search_str)
	#Tsymbol = input_val[Tint]

	for j in range(0,len(input_val)):
		if (search_str in input_val[j]):
			if (Tsymbol == ""):
				Tsymbol = input_val[j]
			else:
				Tsymbol = Tsymbol + "," + input_val[j]	

	return Tsymbol 

def convert_object_to_date(in_df,cur_col,cur_format,output_col=None):
	new_format = []

	for i in range(0,len(in_df)):
		try:
			new_format.append(datetime.strptime(str(in_df.ix[i][cur_col]),cur_format))
		except Exception as e:
			new_format.append(None)
			LOG.info("i = %4d: %s",sys.exc_traceback.tb_lineno,i,e)

	if (output_col is None):
		in_df[cur_col] = new_format
	else:
		in_df[output_col] = new_format
	return in_df

#-----------------------------
# Replace All
# dic format -> dic = {'hello':'goodbye', 'bad':'good', 'yes':'no'} 	
def replace_all(text, dic):
    for i, j in dic.iteritems():
        text = text.replace(i, j)
    return text

def datetime_SF_to_Mysql(datetime):
	return(datetime.replace(".000+0000","").replace("T"," "))

def AttaskDate_to_datetime_NONE(w):
	### Takes in a dataframe and converts the specified column
	if (w != None):
		output = datetime.strptime(w[:-9],'%Y-%m-%dT%H:%M:%S')
	else:
		output = None
	return(output)

def AttaskDate_to_datetime(df,column):
	### Takes in a dataframe and converts the specified column
	df[column] = [datetime.strptime(w[:-9],'%Y-%m-%dT%H:%M:%S') for w in df[column]]
	return(df)

def AttaskDate_to_Mysql(df,column):
	### Takes in a dataframe and converts the specified column
	df[column] = [w.replace('T',' ') for w in df[column]]
	tmp = [w.split(':') for w in df[column]]
	df[column] = [w[0] + ':' + w[1] + ':' + w[2] for w in tmp]
	return(df)

def datetime_to_Mysql(df,column):
	### Takes in a dataframe and converts the specified column
	df[column] = [w.strftime('%Y-%m-%d %H:%M:%S') for w in df[column]]
	return(df)

def convert_Inputdate_to_datetime(val):

	val = val.strip(' ')
	split_val = str(val).split("/")
	
	month = int(split_val[0])
	day = int(split_val[1])

	other_val = split_val[2].split(' ') 
	year = int(other_val[0] )

	hr = int((other_val[1].split(':'))[0] )
	minute = int((other_val[1].split(':'))[1] )
	sec = int((other_val[1].split(':'))[2] )

	#return year.rjust(4,'0') + '-' + month.rjust(2,'0') + '-' + day.rjust(2,'0') + ' ' + hr.rjust(2,'0') + ':' + minute.rjust(2,'0') + ':' + sec.rjust(2,'0')
	return datetime.datetime(year,month,day,hr,minute,sec)

def convert_num_to_month(num):
	if (num == '01'): month = 'Jan'
	if (num == '02'): month = 'Feb'
	if (num == '03'): month = 'Mar'
	if (num == '04'): month = 'Apr'
	if (num == '05'): month = 'May'
	if (num == '06'): month = 'Jun'
	if (num == '07'): month = 'Jul'
	if (num == '08'): month = 'Aug'
	if (num == '09'): month = 'Sep'
	if (num == '10'): month = 'Oct'
	if (num == '11'): month = 'Nov'
	if (num == '12'): month = 'Dec'

	return month

def convertEmployeeBin(val):
	out = None
	if(val < 250):   out = "000000-000249"
	if(val >= 250 and val < 500): out = "000250-000499"
	if(val >= 500 and val < 1000): out = "000500-000999"
	if(val >= 1000 and val < 2500): out = "001000-002499"
	if(val >= 2500 and val < 5000): out = "002500-004999"
	if(val >= 5000 and val < 10000): out = "005000-009999"
	if(val >= 10000 and val < 25000): out = "010000-024999"
	if(val >= 25000 and val < 50000): out = "025000-049999"
	if(val >= 50000 and val < 100000): out = "050000-099999"
	if(val >= 100000): out = "100000+"

#	if(val = '0 and val < 249'):   out = "000000-000249"
#	if(val = '250 - 499'): out = "000249-000499"
#	if(val = '500 - 1000'): out = "000500-001000"
#	if(val = '1001 - 2500'): out = "001001-002499"
#	if(val = '2501 - 5000'): out = "002500-005000"
#	if(val = '5001 - 10000'): out = "005000-009999"
#	if(val = '10001 - 25000'): out = "010000-024999"
#	if(val = '25001 - 50000'): out = "025000-049999"
#	if(val = '50001 - 100000'): out = "050000-099999"
#	if(val = '>100000'): out = "100000+"

	return(out)

def create_sfdc_write_variables():
	############################################
	# Define ONLY write variables for SFDC 
	############################################
	sfdc_write_variables = [] 
	sfdc_write_variables.append('No_of_Videos_Completed__c')
	sfdc_write_variables.append('Percent_Video_Deployment__c')

	return(sfdc_write_variables)

def sfdc_write(sf,sfdc_write_variables,sfdc_account,sfdc_var0,sfdc_var1):

	############################################
	# Check that the write variables are correct
	############################################
	if (len(sfdc_write_variables) > 2):
		LOG.info('Too Many SFDC Write Variables ... %s > 2\nProgram will terminate\n',len(sfdc_write_variables))
		sys.exit()

	if (sfdc_write_variables[0] != 'No_of_Videos_Completed__c'):
		LOG.info('sfdc_write_variable[0] <%s> != <No_of_Videos_Completed__c>\nProgram will terminate\n',sfdc_write_variables[0])
		sys.exit()

	if (sfdc_write_variables[1] != 'Percent_Video_Deployment__c'): 
		LOG.info('sfdc_write_variable[1] <%s> != <Percent_Video_Deployment__c>\nProgram will terminate\n',sfdc_write_variables[1])
		sys.exit()

	for i in range(0,len(sfdc_account)):
		LOG.info('WRITING TO SFDC ACCOUNT %s ... %s ... %s',sfdc_account[i],sfdc_var0[i],sfdc_var1[i])
		sf.Account.update(sfdc_account[i],{sfdc_write_variables[0]:sfdc_var0[i],sfdc_write_variables[1]:sfdc_var1[i]})
	
def createXLSX(file_name,sheet_name,column_order,special_format,data_df,new_output_file):

	LOG.info('Begin Function')

	cur_datetime = datetime.now()
	output_file = file_name + "_" + cur_datetime.strftime('%Y%m%d') +'.xlsx'

	###############################
	# Get column index list
	###############################
	column_idx = []
	for i in range(0,len(column_order)):
		try:
			column_idx.append(list(data_df.columns).index(column_order[i]))
		except:
			LOG.error('<column_order> variable has an incorrect column name')
			sys.exit()


	###############################
	# Create xlsx file
	###############################
	if (new_output_file == True):
		wb = Workbook()
		wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
	else:
		try:
			wb = load_workbook(output_file)
		except:
			wb = Workbook()
			wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

	ws = wb.create_sheet()
	ws.title = sheet_name 

	####################
	# Format xlsx file
	####################
	ws.cell(row = 0, column = 0).style.alignment.wrap_text = True
	#ws.merge_cells('A1:C1')
	for i in range(0,len(column_order)+1):
		ws.cell(row = 0, column = i).style.fill.fill_type = Fill.FILL_SOLID 
		ws.cell(row = 0, column = i).style.fill.start_color.index = Color.DARKBLUE 
		ws.cell(row = 0, column = i).style.font.color = Color(Color.WHITE) 
	for i in range(0,len(column_order)+1):
		ws.cell(row = 0, column = 0).style.alignment.horizontal = "center" 
		ws.cell(row = 0, column = 0).style.alignment.vertical = "center" 

	#########################
	# Create Column Headers
	#########################
	ws.cell(row = 0,column = 0).value = 'unique_id' 
	for i in range(0,len(column_order)):
		ws.cell(row = 0,column = i+1).value = data_df.columns[column_idx[i]] 

	########################
	# Output data 
	########################
	LOG.debug("\n%s",data_df)
	index = data_df.index
	for i in range(0,len(data_df)):
		ws.cell(row = (i+1), column = 0).value = i+1
		for j in range(0,len(column_order)):
			try:
				ws.cell(row = (i+1), column = (j+1)).value = data_df.ix[index[i]][column_order[j]] 
			except Exception as e:
				LOG.error("%s",e)
				ws.cell(row = (i+1), column = (j+1)).value = '' 
			for k in range(0,len(special_format)):
				if (column_order[j] == special_format.keys()[k]):
					ws.cell(row = (i+1), column = (j+1)).style.number_format.format_code = special_format[column_order[j]] 
	
	Nrow = ws.get_highest_row();
	LOG.info('Nrow = %s',Nrow)
	ws.auto_filter = "A1:" + (chr(len(column_order) + ord('A'))) + str(Nrow) # Turn on Autofilter 
	wb.save(output_file)

def createGenericLookup_DF(input_df,title1,title2):

	########################################################
	## Written by DKJ on 6/10/14
	## Creates key-value pair for generic lookup from A[1] = 2 
	########################################################
	LOG.info("A[%s] = %s",title1,title2)

	######################################
	# Find the column for title1 & title2
	######################################
	generic_lookup = {}
	for i in range(0,len(input_df)):
		try:
			generic_lookup[input_df.ix[i][title1]] = input_df.ix[i][title2]
		except Exception as e:
			LOG.error("FAILED . Line %s: %s",sys.exc_traceback.tb_lineno,e)
			generic_lookup[input_df.ix[i][title1]] = None

	return(generic_lookup)

def createGenericLookup(lookup_file,sheet_name,title1,title2):

	########################################################
	## Written by DKJ on 6/10/14
	## Creates key-value pair for generic lookup from A[1] = 2 
	########################################################
	LOG.info("Open :%s: ... ",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name(sheet_name);

	LOG.info("A[%s] = %s",title1,title2)

	######################################
	# Find the column for title1 & title2
	######################################
	for i in range(0,ws.get_highest_column()+1):	
		if (ws.cell(row=0,column=i).value == title1):
			var1 = i
		if (ws.cell(row=0,column=i).value == title2):
			var2 = i

	Nrow = ws.get_highest_row()+1;

	generic_lookup = {}
	for i in range(2,Nrow):
		if (ws.cell(row=(i-1),column=var1).value != None): 
			generic_lookup[ws.cell(row=(i-1),column=var1).value] = ws.cell(row=(i-1),column=var2).value

	return(generic_lookup)

def createCSM_Lookup(lookup_file):

	########################################################
	## Written by DKJ on 9/19/14
	## Creates key-value pair for csm_lookup from sf
	########################################################
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name('g2account_csm_lookup');

	Nrow = ws.get_highest_row()+1;

	sfdc_CSM_lookup = {}
	g2_account_id = []
	g2_account_name = []
	sfdc_account_name = []
	program_start_date = []
	for i in range(2,Nrow):
		g2_account_id.append(ws.cell(row=(i-1),column=0).value)
		g2_account_name.append(ws.cell(row=(i-1),column=1).value)
		sfdc_account_name.append(ws.cell(row=(i-1),column=2).value)
		program_start_date.append(ws.cell(row=(i-1),column=5).value)

	sfdc_CSM_lookup['g2_account_id'] = g2_account_id
	sfdc_CSM_lookup['g2_account_name'] = g2_account_name
	sfdc_CSM_lookup['sfdc_account_name'] = sfdc_account_name
	sfdc_CSM_lookup['program_start_date'] = program_start_date

	return(sfdc_CSM_lookup)

def createG2_CSM_Lookup(lookup_file):

	########################################################
	## Written by DKJ on 9/19/14
	## Creates key-value pair for industry_lookup from sf
	########################################################
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name('csm_lookup');

	Nrow = ws.get_highest_row()+1;

	g2_CSM_lookup = {}
	for i in range(2,Nrow):
		LOG.debug(":%s:%s",ws.cell(row=(i-1),column=0).value,ws.cell(row=(i-1),column=1).value)
		g2_CSM_lookup[ws.cell(row=(i-1),column=0).value] = ws.cell(row=(i-1),column=1).value

	return(g2_CSM_lookup)

def createCSM_g2account_Lookup(lookup_file):

	############################################################
	## Written by DKJ on 9/19/14
	## Creates key-value pair for g2account --> CSM name lookup 
	############################################################
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name('g2account_csm_lookup');

	Nrow = ws.get_highest_row()+1;

	g2account_csm_lookup = {}
	for i in range(2,Nrow):
		g2account_csm_lookup[str(ws.cell(row=(i-1),column=0).value)] = ws.cell(row=(i-1),column=4).value

	return(g2account_csm_lookup)

def create_g2account_SFDCaccount_Lookup(lookup_file):

	############################################################
	## Written by DKJ on 9/19/14
	## Creates key-value pair for g2account --> CSM name lookup 
	############################################################
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name('g2account_csm_lookup');

	Nrow = ws.get_highest_row()+1;

	sfdcaccount_g2account_lookup = {}
	for i in range(2,Nrow):
		sfdcaccount_g2account_lookup[str(ws.cell(row=(i-1),column=2).value)] = ws.cell(row=(i-1),column=0).value

	return(sfdcaccount_g2account_lookup)

def create_SFDCaccount_g2account_Lookup(lookup_file):

	############################################################
	## Written by DKJ on 9/19/14
	## Creates key-value pair for g2account --> CSM name lookup 
	############################################################
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name('g2account_csm_lookup');

	Nrow = ws.get_highest_row()+1;

	g2account_sfdc_lookup = {}
	for i in range(2,Nrow):
		g2account_sfdc_lookup[str(ws.cell(row=(i-1),column=0).value)] = ws.cell(row=(i-1),column=2).value

	return(g2account_sfdc_lookup)

def createIndustryLookup(lookup_file):

	########################################################
	## Written by DKJ on 6/10/14
	## Creates key-value pair for industry_lookup from sf
	########################################################
	#lookup_file = '/media/sf_Data/FY14_analytics/lookup/industry_lookup_20140610.xlsx';
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file);
	ws = wb.get_sheet_by_name('industry_lookup');

	Nrow = ws.get_highest_row()+1;

	industry_lookup = {}
	for i in range(2,Nrow):
		LOG.debug(":%s:%s",ws.cell(row=(i-1),column=0).value,ws.cell(row=(i-1),column=1).value)
		industry_lookup[ws.cell(row=(i-1),column=0).value] = ws.cell(row=(i-1),column=1).value

	return(industry_lookup)

def createNUMIndustryLookup(lookup_file):

	########################################################
	## Written by DKJ on 6/10/14
	## Creates key-value pair for industry_lookup from sf
	########################################################
	#lookup_file = '/media/sf_Data/FY14_analytics/lookup/industry_lookup_20140610.xlsx';
	LOG.info("Open :%s:",lookup_file)

	wb = load_workbook(lookup_file)
	ws = wb.get_sheet_by_name('Nindustry_lookup')

	Nrow = ws.get_highest_row()+1;

	Nindustry_lookup = {}
	for i in range(2,Nrow):
		LOG.debug(":%s:%s",ws.cell(row=(i-1),column=0).value,ws.cell(row=(i-1),column=1).value)
		Nindustry_lookup[ws.cell(row=(i-1),column=0).value] = ws.cell(row=(i-1),column=1).value

	return(Nindustry_lookup)

def createContactLookup(op,contact):

	########################################################
	## Written by DKJ on 6/10/14
	## Creates key-value pair for op.primary_contact --> contactID
	########################################################
	contact_lookup = collections.OrderedDict()
	contact_record = map((lambda item: item['Id']),contact['records']) 
	for x in op['records']:
		if (x['Primary_Contact__c'] != None):
			contact_lookup[x['Primary_Contact__c']] = contact_record.index(x['Primary_Contact__c']) 
			LOG.debug(":%s:%s:%s:",x['Primary_Contact__c'],contact_lookup[x['Primary_Contact__c']],contact_record[contact_lookup[x['Primary_Contact__c']]]) 
		else:
			contact_lookup[x['Primary_Contact__c']] = None 
			
	return(contact_lookup)

def updateIndustryLookup(op_output,lookup_file,industry_lookup):

	industry_list = map((lambda item: item['Account']['Industry']),op_output['records']) 
	sic_list = map((lambda item: item['Account']['Sic']),op_output['records']) 
	naics_list = map((lambda item: item['Account']['NaicsCode']),op_output['records']) 
	wb = load_workbook(lookup_file)
	ws = wb.get_sheet_by_name('industry_lookup')
	Nrow = ws.get_highest_row();

	error = False
	for x in industry_list:
		try:
			tmp = industry_lookup[x]
		except KeyError,e:
			LOG.error("NOTFOUND:%50s:",x)
			LOG.error("%6s: SIC = %10s: NAICS = %10s",industry_list.index(x),sic_list[industry_list.index(x)],naics_list[industry_list.index(x)] )
			ws.cell(row=Nrow,column=0).value = x 
			Nrow = Nrow + 1
			error = True

	wb.save(lookup_file)

	if (error):
		LOG.error("############################################")
		LOG.error("Update lookup file with the names above")
		LOG.error("############################################")

		call(["libreoffice",lookup_file])
		sys.exit(10)

def SicIndustryLookup():

	SIC_industry_lookup = {}
	SIC_industry_lookup["01"] = "Agricultural, Forestry & Fishing"
	SIC_industry_lookup["02"] = "Agricultural, Forestry & Fishing"
	SIC_industry_lookup["07"] = "Agricultural, Forestry & Fishing"
	SIC_industry_lookup["08"] = "Agricultural, Forestry & Fishing"
	SIC_industry_lookup["09"] = "Agricultural, Forestry & Fishing"
	SIC_industry_lookup["10"] = "Mining"
	SIC_industry_lookup["12"] = "Mining"
	SIC_industry_lookup["13"] = "Mining"
	SIC_industry_lookup["14"] = "Mining"
	SIC_industry_lookup["15"] = "Construction"
	SIC_industry_lookup["16"] = "Construction"
	SIC_industry_lookup["17"] = "Construction"
	SIC_industry_lookup["20"] = "Manufacturing"
	SIC_industry_lookup["21"] = "Manufacturing"
	SIC_industry_lookup["22"] = "Manufacturing"
	SIC_industry_lookup["23"] = "Manufacturing"
	SIC_industry_lookup["24"] = "Manufacturing"
	SIC_industry_lookup["25"] = "Manufacturing"
	SIC_industry_lookup["26"] = "Manufacturing"
	SIC_industry_lookup["27"] = "Manufacturing"
	SIC_industry_lookup["28"] = "Manufacturing"
	SIC_industry_lookup["29"] = "Manufacturing"
	SIC_industry_lookup["30"] = "Manufacturing"
	SIC_industry_lookup["31"] = "Manufacturing"
	SIC_industry_lookup["32"] = "Manufacturing"
	SIC_industry_lookup["33"] = "Manufacturing"
	SIC_industry_lookup["34"] = "Manufacturing"
	SIC_industry_lookup["35"] = "Manufacturing"
	SIC_industry_lookup["36"] = "Manufacturing"
	SIC_industry_lookup["37"] = "Manufacturing"
	SIC_industry_lookup["38"] = "Manufacturing"
	SIC_industry_lookup["39"] = "Manufacturing"
	SIC_industry_lookup["40"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["41"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["42"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["43"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["44"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["45"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["46"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["47"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["48"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["49"] = "Transportation, Communications, Electric, Gas & Sanitary"
	SIC_industry_lookup["50"] = "Wholesale Trade"
	SIC_industry_lookup["51"] = "Wholesale Trade"
	SIC_industry_lookup["52"] = "Retail Trade"
	SIC_industry_lookup["53"] = "Retail Trade"
	SIC_industry_lookup["54"] = "Retail Trade"
	SIC_industry_lookup["55"] = "Retail Trade"
	SIC_industry_lookup["56"] = "Retail Trade"
	SIC_industry_lookup["57"] = "Retail Trade"
	SIC_industry_lookup["58"] = "Retail Trade"
	SIC_industry_lookup["59"] = "Retail Trade"
	SIC_industry_lookup["60"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["61"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["62"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["63"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["64"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["65"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["67"] = "Finance, Insurance & Real Estate"
	SIC_industry_lookup["70"] = "Services"
	SIC_industry_lookup["72"] = "Services"
	SIC_industry_lookup["73"] = "Services"
	SIC_industry_lookup["75"] = "Services"
	SIC_industry_lookup["76"] = "Services"
	SIC_industry_lookup["78"] = "Services"
	SIC_industry_lookup["79"] = "Services"
	SIC_industry_lookup["80"] = "Services"
	SIC_industry_lookup["81"] = "Services"
	SIC_industry_lookup["82"] = "Services"
	SIC_industry_lookup["83"] = "Services"
	SIC_industry_lookup["84"] = "Services"
	SIC_industry_lookup["86"] = "Services"
	SIC_industry_lookup["87"] = "Services"
	SIC_industry_lookup["89"] = "Services"
	SIC_industry_lookup["91"] = "Public Administration"
	SIC_industry_lookup["92"] = "Public Administration"
	SIC_industry_lookup["93"] = "Public Administration"
	SIC_industry_lookup["94"] = "Public Administration"
	SIC_industry_lookup["95"] = "Public Administration"
	SIC_industry_lookup["96"] = "Public Administration"
	SIC_industry_lookup["97"] = "Public Administration"
	SIC_industry_lookup["99"] = "Public Administration"

	return(SIC_industry_lookup)

def SicSubIndustryLookup():

	SIC_subindustry_lookup = {}
	SIC_subindustry_lookup["01"] = "Agricultural Production - Crops"
	SIC_subindustry_lookup["02"] = "Agricultural Production - Livestock and Anomal Specialties"
	SIC_subindustry_lookup["07"] = "Agricultural Services"
	SIC_subindustry_lookup["08"] = "Forestry"
	SIC_subindustry_lookup["09"] = "Fishing, Hunting and Trapping"
	SIC_subindustry_lookup["10"] = "Metal Mining"
	SIC_subindustry_lookup["12"] = "Coal Mining"
	SIC_subindustry_lookup["13"] = "Oil and Gas Extraction"
	SIC_subindustry_lookup["14"] = "Mining and Quarrying of Nonmetallic Minerals, Except Fuels"
	SIC_subindustry_lookup["15"] = "Building Cnstrctn - General Contractors & Operative Builders"
	SIC_subindustry_lookup["16"] = "Heavy Cnstrctn, Except Building Construction - Contractors"
	SIC_subindustry_lookup["17"] = "Construction - Special Trade Contractors"
	SIC_subindustry_lookup["20"] = "Food and Kindred Products"
	SIC_subindustry_lookup["21"] = "Tobacco Products"
	SIC_subindustry_lookup["22"] = "Textile Mill Products"
	SIC_subindustry_lookup["23"] = "Apparel, Finished Prdcts from Fabrics & Similar Materials"
	SIC_subindustry_lookup["24"] = "Lumber and Wood Products, Except Furniture"
	SIC_subindustry_lookup["25"] = "Furniture and Fixtures"
	SIC_subindustry_lookup["26"] = "Paper and Allied Products"
	SIC_subindustry_lookup["27"] = "Printing, Publishing and Allied Industries"
	SIC_subindustry_lookup["28"] = "Chemicals and Allied Products"
	SIC_subindustry_lookup["29"] = "Petroleum Refining and Related Industries"
	SIC_subindustry_lookup["30"] = "Rubber and Miscellaneous Plastic Products"
	SIC_subindustry_lookup["31"] = "Leather and Leather Products"
	SIC_subindustry_lookup["32"] = "Stone, Clay, Glass, and Concrete Products"
	SIC_subindustry_lookup["33"] = "Primary Metal Industries"
	SIC_subindustry_lookup["34"] = "Fabricated Metal Prdcts, Except Machinery & Transport Eqpmnt"
	SIC_subindustry_lookup["35"] = "Industrial and Commercial Machinery and Computer Equipment"
	SIC_subindustry_lookup["36"] = "Electronic, Elctrcl Eqpmnt & Cmpnts, Excpt Computer Eqpmnt"
	SIC_subindustry_lookup["37"] = "Transportation Equipment"
	SIC_subindustry_lookup["38"] = "Mesr/Anlyz/Cntrl Instrmnts; Photo/Med/Opt Gds; Watchs/Clocks"
	SIC_subindustry_lookup["39"] = "Miscellaneous Manufacturing Industries"
	SIC_subindustry_lookup["40"] = "Railroad Transportation"
	SIC_subindustry_lookup["41"] = "Local, Suburban Transit & Interurbn Hgwy Passenger Transport"
	SIC_subindustry_lookup["42"] = "Motor Freight Transportation"
	SIC_subindustry_lookup["43"] = "United States Postal Service"
	SIC_subindustry_lookup["44"] = "Water Transportation"
	SIC_subindustry_lookup["45"] = "Transportation by Air"
	SIC_subindustry_lookup["46"] = "Pipelines, Except Natural Gas"
	SIC_subindustry_lookup["47"] = "Transportation Services"
	SIC_subindustry_lookup["48"] = "Communications"
	SIC_subindustry_lookup["49"] = "Electric, Gas and Sanitary Services"
	SIC_subindustry_lookup["50"] = "Wholesale Trade - Durable Goods"
	SIC_subindustry_lookup["51"] = "Wholesale Trade - Nondurable Goods"
	SIC_subindustry_lookup["52"] = "Building Matrials, Hrdwr, Garden Supply & Mobile Home Dealrs"
	SIC_subindustry_lookup["53"] = "General Merchandise Stores"
	SIC_subindustry_lookup["54"] = "Food Stores"
	SIC_subindustry_lookup["55"] = "Automotive Dealers and Gasoline Service Stations"
	SIC_subindustry_lookup["56"] = "Apparel and Accessory Stores"
	SIC_subindustry_lookup["57"] = "Home Furniture, Furnishings and Equipment Stores"
	SIC_subindustry_lookup["58"] = "Eating and Drinking Places"
	SIC_subindustry_lookup["59"] = "Miscellaneous Retail"
	SIC_subindustry_lookup["60"] = "Depository Institutions"
	SIC_subindustry_lookup["61"] = "Nondepository Credit Institutions"
	SIC_subindustry_lookup["62"] = "Security & Commodity Brokers, Dealers, Exchanges & Services"
	SIC_subindustry_lookup["63"] = "Insurance Carriers"
	SIC_subindustry_lookup["64"] = "Insurance Agents, Brokers and Service"
	SIC_subindustry_lookup["65"] = "Real Estate	"
	SIC_subindustry_lookup["67"] = "Holding and Other Investment Offices"
	SIC_subindustry_lookup["70"] = "Hotels, Rooming Houses, Camps, and Other Lodging Places"
	SIC_subindustry_lookup["72"] = "Personal Services"
	SIC_subindustry_lookup["73"] = "Business Services"
	SIC_subindustry_lookup["75"] = "Automotive Repair, Services and Parking"
	SIC_subindustry_lookup["76"] = "Miscellaneous Repair Services"
	SIC_subindustry_lookup["78"] = "Motion Pictures"
	SIC_subindustry_lookup["79"] = "Amusement and Recreation Services"
	SIC_subindustry_lookup["80"] = "Health Services"
	SIC_subindustry_lookup["81"] = "Legal Services"
	SIC_subindustry_lookup["82"] = "Educational Services"
	SIC_subindustry_lookup["83"] = "Social Services	"
	SIC_subindustry_lookup["84"] = "Museums, Art Galleries and Botanical and Zoological Gardens"
	SIC_subindustry_lookup["86"] = "Membership Organizations"
	SIC_subindustry_lookup["87"] = "Engineering, Accounting, Research, Management & Related Svcs"
	SIC_subindustry_lookup["89"] = "Services, Not Elsewhere Classified"
	SIC_subindustry_lookup["91"] = "Executive, Legislative & General Government, Except Finance"
	SIC_subindustry_lookup["92"] = "Justice, Public Order and Safety"
	SIC_subindustry_lookup["93"] = "Public Finance, Taxation and Monetary Policy"
	SIC_subindustry_lookup["94"] = "Administration of Human Resource Programs"
	SIC_subindustry_lookup["95"] = "Administration of Environmental Quality and Housing Programs"
	SIC_subindustry_lookup["96"] = "Administration of Economic Programs"
	SIC_subindustry_lookup["97"] = "National Security and International Affairs"
	SIC_subindustry_lookup["99"] = "Nonclassifiable Establishments"

	return(SIC_subindustry_lookup)

def NaicsIndustryLookup():

	NAICS_industry_lookup = {}
	NAICS_industry_lookup["11"] = "Agricultural, Forestry, Fishing and Hunting"
	NAICS_industry_lookup["21"] = "Mining"
	NAICS_industry_lookup["22"] = "Utilities"
	NAICS_industry_lookup["23"] = "Construction"
	NAICS_industry_lookup["31"] = "Manufacturing"
	NAICS_industry_lookup["32"] = "Manufacturing"
	NAICS_industry_lookup["33"] = "Manufacturing"
	NAICS_industry_lookup["42"] = "Wholesale Trade"
	NAICS_industry_lookup["44"] = "Retail Trade"
	NAICS_industry_lookup["45"] = "Retail Trade"
	NAICS_industry_lookup["48"] = "Transportation and Warehousing"
	NAICS_industry_lookup["49"] = "Transportation and Warehousing"
	NAICS_industry_lookup["51"] = "Information"
	NAICS_industry_lookup["52"] = "Finance and  Insurance"
	NAICS_industry_lookup["53"] = "Real Estate Rental and Leasing"
	NAICS_industry_lookup["54"] = "Professional, Scientific and Technical Services"
	NAICS_industry_lookup["55"] = "Mgmt of Companies and Enterprises"
	NAICS_industry_lookup["56"] = "Administrative and Support, Waste Mgmt and Remediation Services"
	NAICS_industry_lookup["61"] = "Educational Services"
	NAICS_industry_lookup["62"] = "Health Care and Social Assistance"
	NAICS_industry_lookup["71"] = "Arts, Entertainment and Recreation"
	NAICS_industry_lookup["72"] = "Accomodation and Food Services"
	NAICS_industry_lookup["81"] = "Other Services"
	NAICS_industry_lookup["92"] = "Public Administration"

	return(NAICS_industry_lookup)

